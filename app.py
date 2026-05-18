import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

st.set_page_config(page_title="Raventrack ETL Pipeline Dashboard", page_icon="📊", layout="wide")

st.title("📊 Raventrack ETL Pipeline Dashboard")
st.write("Upload your files below to automatically process data and download the cleaned end product.")

# Layout Columns for file uploads
col1, col2 = st.columns(2)

with col1:
    st.subheader("1. Raventrack Export")
    raw_file = st.file_uploader("Upload Raw Raventrack CSV", type=["csv"])

with col2:
    st.subheader("2. Affiliate Log Reference")
    aff_file = st.file_uploader("Upload Data Test File (Excel)", type=["xlsx"])

if raw_file and aff_file:
    if st.button("🚀 Run ETL Pipeline", type="primary"):
        try:
            # --- 1. PROCESS THE CSV RAW FILE ---
            bytes_data = raw_file.getvalue()
            lines = bytes_data.decode('utf-8', errors='ignore').split('\n')
            
            skip_rows = 0
            for i, line in enumerate(lines[:15]):
                if 'player' in line.lower() or 'username' in line.lower():
                    skip_rows = i
                    break
            
            raw_file.seek(0)
            df_raw = pd.read_csv(raw_file, skiprows=skip_rows, float_precision='high', low_memory=False)
            df_raw.columns = df_raw.columns.str.strip()
            
            # Isolate the exact tracking columns your boss's program selected
            columns_to_keep = ['Player Id', 'Affiliate Profile Id', 'Affiliate Profile Username']
            df_clean = df_raw[columns_to_keep].copy()
            df_clean.columns = ['User ID', 'Affiliate Profile ID', 'Affiliate Profile Username']
            
            # --- 2. PROCESS THE EXCEL REFERENCE FILE ---
            df_aff_raw = pd.read_excel(aff_file, sheet_name='RT Tag Reference')
            df_aff_raw.columns = df_aff_raw.columns.astype(str).str.strip()
            
            if 'Affiliate Profile Username' not in df_aff_raw.columns:
                st.error(f"❌ Column Error: Could not find 'Affiliate Profile Username' column in Nicola's file.")
                st.stop()
                
            # --- 3. MATCHING & TRANSFORM LOGIC (EXACTLY LIKE OLD SCRIPT) ---
            df_clean['Affiliate Profile Username'] = df_clean['Affiliate Profile Username'].astype(str).str.strip()
            df_aff_raw['Affiliate Profile Username'] = df_aff_raw['Affiliate Profile Username'].astype(str).str.strip()
            
            # Execute the Left Merge mapping
            df_final = pd.merge(df_clean, df_aff_raw, on='Affiliate Profile Username', how='left')
            
            # Apply exact sort rules from the old script
            df_final = df_final.sort_values(by='Affiliate Profile ID')
            
            # Generate the special integer bulk string format column your boss needs
            df_final['Aardvark Bulk String'] = df_final['User ID'].dropna().astype(float).astype('int64')
            
            # --- 4. SUCCESS & EXPORT ENGINE ---
            st.success("🎉 ETL Pipeline Completed Successfully!")
            st.dataframe(df_final.head(100), use_container_width=True)
            
            # Write to buffer memory and execute styling logic blocks 
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_final.to_excel(writer, index=False, sheet_name='Clean Processed Data')
            
            # Load workbook from memory buffer to apply high-contrast layout formatting rules
            output.seek(0)
            wb = openpyxl.load_workbook(output)
            ws = wb["Clean Processed Data"]
            
            # Apply corporate high-contrast professional style
            ws.views.sheetView[0].showGridLines = True
            ws.freeze_panes = "A2"
            
            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            thin_border = Border(left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
                                 top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB'))
            
            # Render headers
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Render data rows and cell alignment parameters
            row_count = len(df_final)
            for r_idx in range(2, row_count + 2):
                is_even = (r_idx % 2 == 0)
                ws.row_dimensions[r_idx].height = 20
                for c_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = Font(name="Segoe UI", size=11)
                    cell.border = thin_border
                    if is_even and c_idx != ws.max_column:
                        cell.fill = zebra_fill
                    if c_idx in [1, 2, ws.max_column]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto column widths adjustments layout
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = 24
            ws.column_dimensions[get_column_letter(ws.max_column)].width = 32
            
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{row_count + 1}"
            
            # Save out styled workbook payload back into memory download block
            styled_output = io.BytesIO()
            wb.save(styled_output)
            processed_data = styled_output.getvalue()
            
            st.download_button(
                label="📥 Download Cleaned Excel Product",
                data=processed_data,
                file_name="Clean_Processed_Aardvark_Tags.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        except Exception as e:
            st.error(f"❌ An error occurred during processing: {str(e)}")
else:
    st.info("💡 Please upload both required input data files above to initialize processing.")