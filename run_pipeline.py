import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === ⚙️ CONFIGURATION SETTINGS ===
INPUT_CSV_FILE = "./downloads/admin-player-reporting-export-2026-05-13-2026-05-14.csv"
REFERENCE_LOG_FILE = "./Data Test File - Nicolla - May 2026.xlsx" 
EXACT_TAB_NAME = "RT Tag Reference"  
OUTPUT_FOLDER = "./clean_outputs/"   

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def execute_etl_pipeline():
    print(f"\n⚡ [1/4] Loading raw Raventrack CSV: {INPUT_CSV_FILE}")
    target_path = INPUT_CSV_FILE if os.path.exists(INPUT_CSV_FILE) else INPUT_CSV_FILE.replace(".csv", "")
    
    if not os.path.exists(target_path):
        print(f"❌ Error: Cannot find your Raventrack export file inside your downloads folder.")
        return

    try:
        df_raw = pd.read_csv(target_path, low_memory=False)
    except Exception as e:
        print(f"❌ Read Error: Failed to parse your raw export CSV. Details: {e}")
        return

    print("✂️ [2/4] Isolating Player ID, Affiliate Profile ID, and Username...")
    try:
        # Standardize raw headers by stripping hidden spaces
        df_raw.columns = df_raw.columns.str.strip()
        columns_to_keep = ['Player Id', 'Affiliate Profile Id', 'Affiliate Profile Username']
        df_clean = df_raw[columns_to_keep].copy()
        df_clean.columns = ['User ID', 'Affiliate Profile ID', 'Affiliate Profile Username']
    except KeyError as e:
        print(f"❌ Layout Error: Missing required column in raw file. Missing: {e}")
        return

    print(f"🔗 [3/4] Cross-referencing against Master Excel Sheet: '{EXACT_TAB_NAME}'...")
    if not os.path.exists(REFERENCE_LOG_FILE):
        print(f"❌ Error: Reference file missing at root layout: '{REFERENCE_LOG_FILE}'")
        return
        
    try:
        df_aff = pd.read_excel(REFERENCE_LOG_FILE, sheet_name=EXACT_TAB_NAME)
    except Exception as e:
        print(f"❌ Tab Error: Failed to open tab sheet '{EXACT_TAB_NAME}'. Details: {e}")
        return
    
    # FORCE-STRIP ALL EXCEL HEADERS TO ELIMINATE TRAILING SPACES BUGS
    df_aff.columns = df_aff.columns.astype(str).str.strip()
    
    # Check if header matches after stripping spaces
    if 'Affiliate Profile Username' not in df_aff.columns:
        print(f"❌ Column Error: Could not find 'Affiliate Profile Username' header column.")
        print(f"👉 Available columns inside this tab are: {df_aff.columns.tolist()}")
        return

    # Clean data content columns for flawless matching mapping
    df_clean['Affiliate Profile Username'] = df_clean['Affiliate Profile Username'].astype(str).str.strip()
    df_aff['Affiliate Profile Username'] = df_aff['Affiliate Profile Username'].astype(str).str.strip()
    
    df_final = pd.merge(df_clean, df_aff, on='Affiliate Profile Username', how='left')
    df_final = df_final.sort_values(by='Affiliate Profile ID')
    df_final['Aardvark Bulk String'] = df_final['User ID'].dropna().astype(float).astype(int)

    output_path = os.path.join(OUTPUT_FOLDER, "Clean_Processed_Aardvark_Tags.xlsx")
    
    print("🎨 [4/4] Writing and styling final structured high-contrast spreadsheet...")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name="Clean Processed Data", index=False)
        
    style_excel_workbook(output_path, len(df_final))
    print(f"\n🎉 SUCCESS: Data pipeline execution complete!")
    print(f"📁 Grab your final clean corporate Excel sheet here: {output_path}\n")

def style_excel_workbook(file_path, row_count):
    wb = load_workbook(file_path)
    ws = wb["Clean Processed Data"]
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A2" 
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
                         top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB'))
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for r_idx in range(2, row_count + 2):
        is_even = (r_idx % 2 == 0)
        ws.row_dimensions[r_idx].height = 20
        for c_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = Font(name="Segoe UI", size=11)
            cell.border = thin_border
            if is_even and c_idx != ws.max_column:
                cell.fill = zebra_fill
            if c_idx in [1, 2, ws.max_column]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 24
    ws.column_dimensions[get_column_letter(ws.max_column)].width = 32 
        
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{row_count + 1}"
    wb.save(file_path)

if __name__ == "__main__":
    execute_etl_pipeline()